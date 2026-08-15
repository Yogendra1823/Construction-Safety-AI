"""
Seeds realistic demo data so the app is immediately explorable after setup
instead of showing empty dashboards. Safe to re-run — it checks for
existing data first and skips if the database is already populated.

Run directly:  python -m database.seed
Or import seed_demo_data() and call it from the Admin Panel.
"""
import random
from io import BytesIO
from datetime import date, timedelta, datetime

from database.db import session_scope, get_engine
from database.models import (
    Base, User, Project, Milestone, TeamAssignment, Document, MaterialItem,
    BudgetItem, Expense, Worker, Attendance, Equipment, InventoryItem,
    PurchaseOrder, SafetyIncident, SafetyChecklistItem, EmergencyContact,
    Notification, ActivityLog, AIRecommendation, ProgressLog,
)
from auth.auth_utils import hash_password
from ai_engine.material_estimator import estimate_materials
from ai_engine.document_ai import _heuristic_category
from utils.files import guess_mime_type

random.seed(7)

PROJECT_SEEDS = [
    dict(project_code="SKY-2026", name="Skyline Financial Tower Phase 2", client_name="Meridian Capital Group",
         location="Downtown Financial District, Sector 4", building_type="Commercial Complex",
         construction_style="Modern", plot_size_sqft=18000, floors=22, bedrooms=0, bathrooms=44,
         kitchens=2, parking_spots=120, quality="Premium", status="Ongoing", progress=68,
         budget_total=61_700_000, spend_pct=0.64, start_offset=-260, duration=560),
    dict(project_code="MTR-2026", name="Metro Hub Underground Terminal", client_name="City Transit Authority",
         location="Central Station Transit Corridor", building_type="Commercial Complex",
         construction_style="Industrial", plot_size_sqft=42000, floors=3, bedrooms=0, bathrooms=18,
         kitchens=1, parking_spots=0, quality="Standard", status="Ongoing", progress=84,
         budget_total=38_400_000, spend_pct=0.79, start_offset=-400, duration=480),
    dict(project_code="PAC-BRG", name="Pacific Waterfront Bridge Expansion", client_name="Dept. of Public Infrastructure",
         location="North Bay Estuary Crossing", building_type="Commercial Complex",
         construction_style="Industrial", plot_size_sqft=9000, floors=1, bedrooms=0, bathrooms=2,
         kitchens=0, parking_spots=0, quality="Premium", status="Delayed", progress=41,
         budget_total=54_200_000, spend_pct=0.58, start_offset=-300, duration=420),
    dict(project_code="ECO-R80", name="Biotech Campus R&D Facility", client_name="Helix Biosciences Inc.",
         location="Innovation Park, Lot 14", building_type="Office",
         construction_style="Contemporary", plot_size_sqft=15000, floors=6, bedrooms=0, bathrooms=24,
         kitchens=4, parking_spots=80, quality="Premium", status="Ongoing", progress=92,
         budget_total=29_800_000, spend_pct=0.90, start_offset=-330, duration=380),
    dict(project_code="SKV-114", name="Skyview Residency Block C", client_name="Nandan Developers",
         location="Whitefield Extension", building_type="Apartment",
         construction_style="Contemporary", plot_size_sqft=6200, floors=9, bedrooms=36, bathrooms=40,
         kitchens=9, parking_spots=45, quality="Standard", status="Completed", progress=100,
         budget_total=8_900_000, spend_pct=0.97, start_offset=-620, duration=520),
    dict(project_code="GRV-002", name="Greenview Villa", client_name="Mr. & Mrs. Kapoor",
         location="Lakeside Colony", building_type="Villa",
         construction_style="Minimalist", plot_size_sqft=3200, floors=2, bedrooms=4, bathrooms=5,
         kitchens=1, parking_spots=2, quality="Premium", status="Planning", progress=6,
         budget_total=1_450_000, spend_pct=0.08, start_offset=-20, duration=300),
    dict(project_code="HLC-990", name="Hillcrest Public School Block", client_name="State Education Board",
         location="Hillcrest Township", building_type="School",
         construction_style="Traditional", plot_size_sqft=11000, floors=3, bedrooms=0, bathrooms=16,
         kitchens=1, parking_spots=20, quality="Basic", status="On Hold", progress=27,
         budget_total=5_250_000, spend_pct=0.30, start_offset=-180, duration=360),
]

WORKER_NAMES = [
    "Arun Mehta", "Fatima Sheikh", "Ravi Kumar", "Priya Nair", "John D'Souza",
    "Suresh Reddy", "Anjali Verma", "Vikram Singh", "Meera Iyer", "Karan Malhotra",
    "Divya Krishnan", "Imran Qureshi", "Neha Joshi", "Rahul Ghosh", "Sneha Pillai",
    "Manoj Yadav", "Pooja Chawla", "Sanjay Bhatt", "Lakshmi Rao", "Tariq Ahmed",
]

MATERIAL_CATEGORY_STATUS = ["In Stock", "In Stock", "In Stock", "Low", "Out of Stock"]


def _placeholder_file_bytes(filename: str, project_name: str) -> bytes:
    """Generates a real, valid file for demo documents — not just metadata —
    so every download in a fresh demo actually opens correctly."""
    ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""
    title = filename.rsplit(".", 1)[0]

    if ext == "pdf":
        from reportlab.lib.pagesizes import A4
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.lib.units import mm
        buf = BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=A4, topMargin=25 * mm)
        styles = getSampleStyleSheet()
        doc.build([
            Paragraph(title, styles["Title"]),
            Spacer(1, 8),
            Paragraph(f"Project: {project_name}", styles["Normal"]),
            Spacer(1, 12),
            Paragraph(
                "This is placeholder demo content generated by the Construction Intelligence "
                "Hub seed script, so the document library is immediately explorable with real, "
                "openable files. Replace it by uploading your own document from the Documents page.",
                styles["Normal"],
            ),
        ])
        return buf.getvalue()

    if ext == "xlsx":
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws["A1"] = title
        ws["A2"] = f"Project: {project_name}"
        ws["A3"] = "Placeholder demo content generated by the CI Hub seed script."
        ws.column_dimensions["A"].width = 60
        buf = BytesIO()
        wb.save(buf)
        return buf.getvalue()

    return f"{title}\n\nProject: {project_name}\n\nPlaceholder demo content generated by the CI Hub seed script.".encode("utf-8")


def _already_seeded(db) -> bool:
    return db.query(Project).count() > 0


def seed_demo_data(force: bool = False):
    engine = get_engine()
    if force:
        Base.metadata.drop_all(engine)
    Base.metadata.create_all(engine)

    with session_scope() as db:
        if _already_seeded(db) and not force:
            return False, "Demo data already present — skipped."

        # --- Users ---
        admin = User(full_name="Admin User", email="admin@cihub.com",
                     password_hash=hash_password("Admin@123"), role="Project Manager",
                     phone="+1 (555) 123-4567", department="Construction Operations", is_admin=True)
        pm2 = User(full_name="Kavya Porwal", email="kavya@cihub.com",
                   password_hash=hash_password("Demo@123"), role="Site Engineer", department="Site Ops")
        pm3 = User(full_name="Rajat Sharma", email="rajat@cihub.com",
                   password_hash=hash_password("Demo@123"), role="Contractor", department="Execution")
        db.add_all([admin, pm2, pm3])
        db.flush()

        users = [admin, pm2, pm3]

        for seed in PROJECT_SEEDS:
            start = date.today() + timedelta(days=seed["start_offset"])
            expected = start + timedelta(days=seed["duration"])
            budget_used = round(seed["budget_total"] * seed["spend_pct"], 2)

            project = Project(
                project_code=seed["project_code"], name=seed["name"], client_name=seed["client_name"],
                location=seed["location"], building_type=seed["building_type"],
                construction_style=seed["construction_style"], plot_size_sqft=seed["plot_size_sqft"],
                plot_length_ft=round((seed["plot_size_sqft"]) ** 0.5, 1),
                plot_width_ft=round((seed["plot_size_sqft"]) ** 0.5, 1),
                floors=seed["floors"], bedrooms=seed["bedrooms"], bathrooms=seed["bathrooms"],
                kitchens=seed["kitchens"], parking_spots=seed["parking_spots"],
                material_quality=seed["quality"], budget_total=seed["budget_total"],
                budget_used=budget_used, start_date=start, expected_completion=expected,
                actual_completion=expected if seed["status"] == "Completed" else None,
                status=seed["status"], progress_percent=seed["progress"],
                risk_score=round(random.uniform(8, 35) if seed["status"] not in ("Delayed", "On Hold") else random.uniform(45, 82), 1),
                created_by=admin.id,
            )
            db.add(project)
            db.flush()

            # Milestones
            phase_names = ["Site Mobilization", "Foundation", "Structural Frame", "MEP Rough-in", "Finishing", "Handover"]
            for i, ph in enumerate(phase_names):
                due = start + timedelta(days=int(seed["duration"] * (i + 1) / len(phase_names)))
                status = "Done" if seed["progress"] > (i + 1) / len(phase_names) * 100 else (
                    "In Progress" if seed["progress"] > i / len(phase_names) * 100 else "Pending")
                if seed["status"] == "Delayed" and status == "In Progress":
                    status = "Delayed"
                db.add(Milestone(project_id=project.id, title=ph, due_date=due,
                                  completed_date=due if status == "Done" else None, status=status))

            # Team assignment
            for u in random.sample(users, k=min(2, len(users))):
                db.add(TeamAssignment(project_id=project.id, user_id=u.id,
                                       role_on_project=random.choice(["Site Supervisor", "Project Lead", "QA Engineer"])))

            # Documents — generate real, valid file bytes (not just metadata) so every
            # download in the demo actually opens correctly.
            doc_names = ["Site Layout Plan.pdf", "Structural Drawings.pdf", "Client Contract.pdf",
                         "Material Purchase Order.xlsx", "Safety Compliance Cert.pdf"]
            for dn in random.sample(doc_names, k=3):
                content = _placeholder_file_bytes(dn, project.name)
                category = _heuristic_category(dn)
                db.add(Document(project_id=project.id, name=dn, category=category, category_source="manual",
                                 size_kb=round(len(content) / 1024, 1), status="ready", content=content,
                                 mime_type=guess_mime_type(dn),
                                 uploaded_by=admin.id, uploaded_at=datetime.utcnow() - timedelta(days=random.randint(1, 60))))

            # Materials — reuse the AI estimator so demo numbers are consistent with the engine
            estimated = estimate_materials(seed["plot_size_sqft"], seed["floors"], seed["quality"])
            for mat in estimated:
                used_ratio = min(seed["progress"] / 100 * random.uniform(0.85, 1.1), 1.0)
                used_qty = round(mat["estimated_qty"] * used_ratio, 1)
                availability = random.choices(MATERIAL_CATEGORY_STATUS, weights=[45, 25, 15, 10, 5])[0]
                db.add(MaterialItem(
                    project_id=project.id, material_name=mat["material_name"], category=mat["category"],
                    unit=mat["unit"], estimated_qty=mat["estimated_qty"], used_qty=used_qty,
                    unit_cost=mat["unit_cost"], supplier=random.choice(["BuildRight Supplies", "Apex Materials Co.", "Metro Hardware", "Prime Cement & Steel"]),
                    availability=availability,
                ))

            # Budget items
            splits = {"Materials": 0.35, "Labor": 0.25, "Equipment": 0.15, "Subcontractors": 0.12, "Overhead": 0.08, "Contingency": 0.05}
            for cat, pct in splits.items():
                allocated = round(seed["budget_total"] * pct, 2)
                spent = round(allocated * seed["spend_pct"] * random.uniform(0.85, 1.15), 2)
                db.add(BudgetItem(project_id=project.id, category=cat, allocated_amount=allocated, spent_amount=min(spent, allocated * 1.3)))

            for _ in range(4):
                db.add(Expense(project_id=project.id, category=random.choice(list(splits.keys())),
                                description=random.choice(["Material delivery", "Crew payroll", "Equipment rental", "Site utilities"]),
                                amount=round(random.uniform(5000, 120000), 2),
                                expense_date=date.today() - timedelta(days=random.randint(1, 45)),
                                recorded_by=admin.id))

            # Workers + attendance
            n_workers = max(4, min(14, int(seed["plot_size_sqft"] / 1500)))
            project_workers = []
            for name in random.sample(WORKER_NAMES, k=min(n_workers, len(WORKER_NAMES))):
                w = Worker(name=name, role=random.choice(
                    ["Site Engineer", "Mason", "Electrician", "Plumber", "Painter", "Labourer", "Carpenter"]),
                    phone=f"+1 (555) {random.randint(200,999)}-{random.randint(1000,9999)}",
                    daily_wage=round(random.uniform(900, 3200), 0), status="Active", project_id=project.id)
                db.add(w)
                db.flush()
                project_workers.append(w)
                for d in range(7):
                    day = date.today() - timedelta(days=d)
                    status = random.choices(["Present", "Absent", "Half-day"], weights=[82, 8, 10])[0]
                    db.add(Attendance(worker_id=w.id, project_id=project.id, work_date=day,
                                       status=status, hours_worked=8 if status == "Present" else (4 if status == "Half-day" else 0)))

            # Equipment
            for _ in range(random.randint(1, 3)):
                db.add(Equipment(name=f"{random.choice(['Tower Crane', 'Concrete Mixer', 'Dump Truck', 'Excavator', 'Generator'])} #{random.randint(100,999)}",
                                  equipment_type=random.choice(["Crane", "Concrete Mixer", "Truck", "Excavator", "Generator"]),
                                  project_id=project.id, status=random.choice(["Available", "In Use", "In Use", "Maintenance"]),
                                  fuel_level=round(random.uniform(20, 100), 0), usage_hours=round(random.uniform(50, 4000), 0),
                                  last_maintenance=date.today() - timedelta(days=random.randint(5, 90)),
                                  next_maintenance=date.today() + timedelta(days=random.randint(5, 60))))

            # Safety
            if random.random() < 0.5:
                db.add(SafetyIncident(project_id=project.id,
                                       incident_type=random.choice(["Minor slip", "PPE violation", "Equipment near-miss", "Electrical hazard"]),
                                       severity=random.choice(["Low", "Medium", "High"]),
                                       description="Logged during routine site walk-through.",
                                       incident_date=date.today() - timedelta(days=random.randint(1, 30)),
                                       status=random.choice(["Open", "Resolved"]), reported_by=admin.id))
            for item in ["Hard Hats", "Safety Harnesses", "Fire Extinguishers", "Scaffolding Inspection", "First Aid Kits"]:
                db.add(SafetyChecklistItem(project_id=project.id, item_name=item,
                                            category="PPE" if "Hat" in item or "Harness" in item else "Site",
                                            status=random.choices(["Pass", "Fail", "Pending"], weights=[75, 10, 15])[0],
                                            checked_date=date.today() - timedelta(days=random.randint(0, 14))))

            # Notifications tied to this project are now generated dynamically in the UI
            
            db.add(ActivityLog(user_id=admin.id, project_id=project.id, action=f"Project '{project.name}' status set to {seed['status']}"))

            # Weekly progress history (last 12 weeks) trending toward current progress
            weeks = 12
            final_actual = seed["progress"]
            elapsed_frac = min(max((date.today() - start).days / max(seed["duration"], 1), 0), 1)
            final_planned = round(elapsed_frac * 100, 1)
            for w in range(weeks, 0, -1):
                log_date = date.today() - timedelta(weeks=w - 1)
                frac = (weeks - w + 1) / weeks
                actual = max(0, round(final_actual * frac * random.uniform(0.9, 1.05), 1))
                planned = max(0, round(final_planned * frac, 1))
                db.add(ProgressLog(project_id=project.id, log_date=log_date,
                                    actual_progress_pct=min(actual, 100), planned_progress_pct=min(planned, 100)))

        # Global inventory + suppliers + purchase orders (not tied to one project)
        for item, cat, qty, unit, reorder in [
            ("Portland Cement", "Structural", 4200, "bags", 800),
            ("TMT Steel Bars", "Structural", 18500, "kg", 4000),
            ("River Sand", "Structural", 3100, "cft", 600),
            ("Vitrified Tiles", "Finishing", 9200, "sqft", 1500),
            ("Exterior Emulsion Paint", "Finishing", 640, "litres", 150),
            ("PVC Conduit Pipes", "Electrical", 2100, "pieces", 400),
        ]:
            db.add(InventoryItem(item_name=item, category=cat, quantity_in_stock=qty, unit=unit,
                                  reorder_level=reorder, warehouse_location=random.choice(["Warehouse A", "Warehouse B", "Site Yard 2"])))

        for _ in range(6):
            db.add(PurchaseOrder(project_id=random.choice([None, None, 1, 2, 3]),
                                  supplier_name=random.choice(["BuildRight Supplies", "Apex Materials Co.", "Metro Hardware", "Prime Cement & Steel"]),
                                  item_name=random.choice(["Cement", "Steel Rebar", "Sand", "Tiles", "Paint"]),
                                  quantity=round(random.uniform(200, 3000), 0), unit_cost=round(random.uniform(50, 500), 2),
                                  order_date=date.today() - timedelta(days=random.randint(1, 20)),
                                  expected_delivery=date.today() + timedelta(days=random.randint(2, 15)),
                                  status=random.choice(["Pending", "Shipped", "Delivered"])))

        for name, role, phone, ctype in [
            ("City Fire Department", "Emergency Response", "911", "Fire"),
            ("Dr. Alicia Fernandes", "Site Medical Officer", "+1 (555) 300-1122", "Medical"),
            ("Metro Police - Sector 4", "Law Enforcement", "911", "Police"),
            ("Ben Okafor", "Site Safety Officer", "+1 (555) 400-7788", "Site Safety"),
        ]:
            db.add(EmergencyContact(name=name, role=role, phone=phone, contact_type=ctype))

        return True, "Demo data seeded successfully."


if __name__ == "__main__":
    ok, msg = seed_demo_data(force=True)
    print(msg)
